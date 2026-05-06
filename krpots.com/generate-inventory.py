import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import io, os

MEDIA     = r"c:\Development\aucdt-utilities\krpots.com\public\media\pots-by-kr"
OUT       = r"c:\Development\aucdt-utilities\krpots.com\krpots-inventory.xlsx"
PROD_BASE = "https://krpots.com/media/pots-by-kr"

# (img_id, title, category, technique, status, price_usd, description, sku)
pieces = [
  ("IMG_1058","Stoneware Form No. 1","Vessels & Vases","Wheel-thrown","For Sale",50.00,"A considered wheel-thrown form with quiet surface variation under a semi-matte glaze.","KR-VV-010"),
  ("IMG_1060","Stoneware Form No. 2","Vessels & Vases","Wheel-thrown","For Sale",50.00,"Companion piece in the stoneware form series. A balanced profile with a confident shoulder.","KR-VV-011"),
  ("IMG_1063","Glazed Vessel Study","Vessels & Vases","Wheel-thrown","For Sale",50.00,"A glaze study vessel showing the interaction of two overlapping glazes at the shoulder line.","KR-VV-012"),
  ("IMG_1064","Iron Wash Vessel","Vessels & Vases","Wheel-thrown","For Sale",50.00,"Iron wash applied over a pale base glaze creates warm, atmospheric depth across the surface.","KR-VV-013"),
  ("IMG_1065","Bottle Form No. 1","Vessels & Vases","Wheel-thrown","For Sale",50.00,"A slender bottle form with a narrow neck and swelling body.","KR-VV-014"),
  ("IMG_1066","Bottle Form No. 2","Vessels & Vases","Wheel-thrown","For Sale",50.00,"Second bottle in the series, with a broader shoulder and subtler neck transition.","KR-VV-015"),
  ("IMG_1067","Raw Clay Vessel","Vessels & Vases","Wheel-thrown","For Sale",50.00,"Minimal glaze intervention lets the fired stoneware body speak for itself.","KR-VV-016"),
  ("IMG_1068","Soda Fired Vessel","Vessels & Vases","Wheel-thrown","For Sale",50.00,"Soda firing produces a distinctive flashing across the surface.","KR-VV-017"),
  ("IMG_1072","Footed Bowl No. 1","Bowls","Wheel-thrown","For Sale",50.00,"A trimmed foot ring elevates this bowl with studied elegance.","KR-BW-011"),
  ("IMG_1074","Footed Bowl No. 2","Bowls","Wheel-thrown","For Sale",50.00,"Companion to Footed Bowl No. 1 with a slightly wider form and deeper glaze colour.","KR-BW-012"),
  ("IMG_1075","Carved Rim Bowl","Bowls","Wheel-thrown","For Sale",50.00,"Carved facets along the exterior create a rhythmic pattern that catches raking light.","KR-BW-013"),
  ("IMG_2544-EDIT","Editorial Study No. 1","Studio & Exhibition","Wheel-thrown","For Sale",50.00,"Professionally photographed studio work showing form and glaze in editorial context.","KR-SE-015"),
  ("IMG_2547-EDIT","Editorial Study No. 2","Studio & Exhibition","Wheel-thrown","For Sale",50.00,"Second editorial study with controlled lighting emphasising surface texture and form.","KR-SE-016"),
  ("IMG_2548","Thrown & Altered Form","Vessels & Vases","Wheel-thrown","For Sale",50.00,"A wheel-thrown form altered after throwing to introduce an asymmetric shoulder profile.","KR-VV-018"),
  ("IMG_2549","Layered Glaze Vase","Vessels & Vases","Wheel-thrown","For Sale",50.00,"Three overlapping glaze layers create depth and movement across this tall vase form.","KR-VV-019"),
  ("IMG_2550-EDIT","Editorial Study No. 3","Studio & Exhibition","Slab-built","For Sale",50.00,"Slab-built form captured in editorial photography.","KR-SE-017"),
  ("IMG_2551","Iron Red Vessel","Vessels & Vases","Wheel-thrown","For Sale",50.00,"Iron red glaze in reduction produces a deep, lustrous surface with flame-driven variation.","KR-VV-020"),
  ("IMG_2554-EDIT","Editorial Study No. 4","Studio & Exhibition","Wheel-thrown","For Sale",50.00,"Fourth editorial study emphasising the dialogue between form and negative space.","KR-SE-018"),
  ("IMG_2556-EDIT","Editorial Study No. 5","Studio & Exhibition","Hand-built","For Sale",50.00,"Hand-built form in editorial context, surface texture prominent under directional lighting.","KR-SE-019"),
  ("IMG_2558","Shino Glaze Pot","Vessels & Vases","Wheel-thrown","For Sale",50.00,"Classic Shino glaze with characteristic orange scorch marks and a crawled milky surface.","KR-VV-021"),
  ("IMG_2559-EDIT","Editorial Study No. 6","Studio & Exhibition","Wheel-thrown","For Sale",50.00,"Sixth editorial study, focusing on the profile silhouette and rim detail of a tall vase.","KR-SE-020"),
  ("IMG_2560","Pale Celadon Vase","Vessels & Vases","Wheel-thrown","For Sale",50.00,"Delicate pale celadon over a refined thrown form.","KR-VV-022"),
  ("IMG_2562","Thrown Mug with Handle","Mugs & Cups","Wheel-thrown","For Sale",50.00,"A well-proportioned mug with a pulled handle that integrates naturally into the thrown form.","KR-MC-010"),
  ("IMG_2563","Yunomi Cup","Mugs & Cups","Wheel-thrown","For Sale",50.00,"A Japanese-influenced yunomi form: tall, handleless, and perfectly weighted for two-handed use.","KR-MC-011"),
  ("IMG_2564","Faceted Cup","Mugs & Cups","Wheel-thrown","For Sale",50.00,"Post-throwing faceting transforms the round form into an angular vessel with strong visual rhythm.","KR-MC-012"),
  ("IMG_2566-EDIT","Editorial Study No. 7","Studio & Exhibition","Coil-built","For Sale",50.00,"Coil-built form in editorial photography, the coil texture intentionally preserved.","KR-SE-021"),
  ("IMG_2568","Copper Glaze Bowl","Bowls","Wheel-thrown","For Sale",50.00,"Copper carbonate in reduction produces a shifting green-red-bronze glaze surface.","KR-BW-014"),
  ("IMG_2569","Ribbed Serving Bowl","Bowls","Wheel-thrown","For Sale",50.00,"Vertical ribbing applied on the wheel gives this serving bowl a textured exterior.","KR-BW-015"),
  ("IMG_2570","Open Form Bowl","Bowls","Wheel-thrown","For Sale",50.00,"A wide, open bowl form with a gently undulating rim and a smooth celadon interior.","KR-BW-016"),
  ("IMG_2571","Slip Decorated Bowl","Bowls","Wheel-thrown","For Sale",50.00,"White slip applied in trailing bands over a dark stoneware body.","KR-BW-017"),
  ("IMG_2573","Black Stoneware Jug","Pitchers & Jugs","Wheel-thrown","For Sale",50.00,"Matte black stoneware jug with a confident, upright form and a cleanly pulled spout.","KR-PJ-007"),
  ("IMG_2574","Speckled Pitcher","Pitchers & Jugs","Wheel-thrown","For Sale",50.00,"Iron speckle through a cream stoneware glaze gives this pitcher a warm, freckled character.","KR-PJ-008"),
  ("IMG_2576","Low Platter No. 1","Platters & Dishes","Wheel-thrown","For Sale",50.00,"A low, wide platter thrown on the wheel with a broad flat floor and a refined trimmed foot.","KR-PD-007"),
  ("IMG_2578-EDIT","Editorial Study No. 8","Studio & Exhibition","Wheel-thrown","For Sale",50.00,"Eighth editorial study, a close study of a rim detail and glaze break.","KR-SE-022"),
  ("IMG_2579","Salt Fired Jug","Pitchers & Jugs","Wheel-thrown","For Sale",50.00,"Salt firing gives this jug a textured orange-peel surface and a natural, unplanned character.","KR-PJ-009"),
  ("IMG_2583","Carved Vessel No. 1","Vessels & Vases","Wheel-thrown","For Sale",50.00,"Carved facets cut into the leather-hard clay create a geometric pattern beneath the glaze.","KR-VV-023"),
  ("IMG_2584","Carved Vessel No. 2","Vessels & Vases","Wheel-thrown","For Sale",50.00,"Second carved vessel with a finer incision pattern and a contrasting interior glaze.","KR-VV-024"),
  ("IMG_2586","Tea Bowl No. 1","Bowls","Hand-built","For Sale",50.00,"A hand-built chawan form with an irregular rim and a pooled glaze interior. Made for ceremony.","KR-BW-018"),
  ("IMG_2587","Tea Bowl No. 2","Bowls","Hand-built","For Sale",50.00,"Companion chawan with a warmer glaze tone and a slightly broader foot placement.","KR-BW-019"),
  ("IMG_2588","Low Platter No. 2","Platters & Dishes","Wheel-thrown","For Sale",50.00,"Second in the low platter series, with a broader floor and a more pronounced rim profile.","KR-PD-008"),
  ("IMG_8521","Studio Piece No. 1","Vessels & Vases","Wheel-thrown","For Sale",50.00,"A studio piece from a recent firing cycle, showing refined control of form and glaze.","KR-VV-025"),
  ("IMG_8639","Studio Piece No. 2","Vessels & Vases","Wheel-thrown","For Sale",50.00,"Companion studio piece with a distinctive glaze break at the shoulder.","KR-VV-026"),
  ("IMG_8659","Studio Piece No. 3","Vessels & Vases","Wheel-thrown","For Sale",50.00,"Third studio piece, notable for its upright profile and a clean, restrained glaze surface.","KR-VV-027"),
  ("IMG_0150","Studio Survey I","Studio & Exhibition","Wheel-thrown","Private Collection",None,"A studio documentation piece capturing an early survey of forms. Preserved for archive reference.","KR-SE-001"),
  ("IMG_0171","Exhibition Study I","Studio & Exhibition","Hand-built","Private Collection",None,"Recorded during a group exhibition installation. Part of the permanent studio archive.","KR-SE-002"),
  ("IMG_0173","Exhibition Study II","Studio & Exhibition","Wheel-thrown","Private Collection",None,"A companion study from the same exhibition series, demonstrating glaze development over time.","KR-SE-003"),
  ("IMG_0176","Studio Survey II","Studio & Exhibition","Hand-built","Private Collection",None,"Survey photograph from the mid-period studio, showing shelf arrangements and form relationships.","KR-SE-004"),
  ("IMG_0178","Archive Document I","Studio & Exhibition","Slab-built","Private Collection",None,"Primary archive document from a decade of kiln experiments. Foundational to the body of work.","KR-SE-005"),
  ("IMG_0179","Archive Document II","Studio & Exhibition","Wheel-thrown","Private Collection",None,"Continuation of the archive documentation series. Surfaces record decades of fire and glaze.","KR-SE-006"),
  ("IMG_0180","Studio Survey III","Studio & Exhibition","Coil-built","Private Collection",None,"Third survey from the studio documentation project, focusing on coil-built structural forms.","KR-SE-007"),
  ("IMG_0182","Exhibition Record I","Studio & Exhibition","Hand-built","Private Collection",None,"Exhibition record from a solo show. Captures the spatial relationship between grouped works.","KR-SE-008"),
  ("IMG_0183","Exhibition Record II","Studio & Exhibition","Wheel-thrown","Private Collection",None,"Second exhibition record from the same installation. The light reveals each form's presence.","KR-SE-009"),
  ("IMG_0337","Archive Document III","Studio & Exhibition","Slab-built","Private Collection",None,"A mid-archive document showing the evolution of slab construction in the studio practice.","KR-SE-010"),
  ("IMG_0356","Archive Document IV","Studio & Exhibition","Wheel-thrown","Private Collection",None,"Fourth archive document in the series, notable for its exploration of reduction fire glazes.","KR-SE-011"),
  ("IMG_0940","Studio Record I","Studio & Exhibition","Hand-built","Private Collection",None,"A candid studio record showing the working environment and process behind the collection.","KR-SE-012"),
  ("IMG_1016","Studio Record II","Studio & Exhibition","Wheel-thrown","Private Collection",None,"Companion to Studio Record I. Together they form a portrait of the studio's character.","KR-SE-013"),
  ("IMG_5180","Exhibition Survey III","Studio & Exhibition","Coil-built","Private Collection",None,"Recent exhibition survey documenting the most current body of work in assembled context.","KR-SE-014"),
  ("IMG_5017","Iron Slip Mug","Mugs & Cups","Wheel-thrown","For Sale",50.00,"Iron slip trails over a warm stoneware body, pooling softly at the base. Holds heat long after the pour.","KR-MC-001"),
  ("IMG_5018","Celadon Cup","Mugs & Cups","Wheel-thrown","For Sale",50.00,"Pale celadon glaze settles into the thrown lines with quiet precision. A cup made for deliberate mornings.","KR-MC-002"),
  ("IMG_5041","Cobalt Stripe Mug","Mugs & Cups","Wheel-thrown","For Sale",50.00,"Bold cobalt banding encircles a well-balanced mug form.","KR-MC-003"),
  ("IMG_5042","Ash Glaze Mug","Mugs & Cups","Wheel-thrown","For Sale",50.00,"Wood ash glaze gives this mug its irregular, lichen-like surface. No two firings produce the same result.","KR-MC-004"),
  ("IMG_5080","Tenmoku Mug","Mugs & Cups","Wheel-thrown","For Sale",50.00,"Deep tenmoku glaze breaks to a rust hue along every thrown ridge.","KR-MC-005"),
  ("IMG_5081","Oatmeal Glaze Mug","Mugs & Cups","Wheel-thrown","For Sale",50.00,"Matte oatmeal glaze lends a quiet, tactile surface that rewards daily use.","KR-MC-006"),
  ("IMG_5091","Textured Cylinder Cup","Mugs & Cups","Hand-built","For Sale",50.00,"Hand-built cylinder with impressed texture along the lower register.","KR-MC-007"),
  ("IMG_5181","Salt Glaze Mug","Mugs & Cups","Wheel-thrown","For Sale",50.00,"Salt-fired stoneware with an orange-peel surface texture unique to this ancient firing method.","KR-MC-008"),
  ("IMG_5282","Brushwork Mug","Mugs & Cups","Wheel-thrown","For Sale",50.00,"Loose brushwork in iron oxide decorates the exterior with gestural strokes beneath a clear glaze.","KR-MC-009"),
  ("IMG_4931","Flared Rim Bowl","Bowls","Wheel-thrown","For Sale",50.00,"A wide-thrown bowl with a gently flared rim and a luminous celadon interior.","KR-BW-001"),
  ("IMG_4969","Iron Oxide Serving Bowl","Bowls","Wheel-thrown","For Sale",50.00,"Generous proportions and a rich iron oxide exterior make this bowl a presence on any table.","KR-BW-002"),
  ("IMG_5214","Wax Resist Bowl","Bowls","Wheel-thrown","For Sale",50.00,"Wax-resist decoration creates a soft pattern where raw clay shows through a layered glaze surface.","KR-BW-003"),
  ("IMG_5215","Celadon Footed Bowl","Bowls","Wheel-thrown","Private Collection",None,"A delicately footed bowl with a pooled celadon glaze. Retained as a touchstone of the glaze series.","KR-BW-004"),
  ("IMG_5229","Nesting Bowl No. 1","Bowls","Wheel-thrown","For Sale",50.00,"The first in a nesting set, thrown to stack cleanly. Ash glaze with a pale blue-green center.","KR-BW-005"),
  ("IMG_5248","Deep Well Bowl","Bowls","Wheel-thrown","For Sale",50.00,"Steep walls and a narrow base give this bowl an unusual depth. Ideal for ramen, soup, or contemplation.","KR-BW-006"),
  ("IMG_5285","Crackle Glaze Bowl","Bowls","Wheel-thrown","Private Collection",None,"Fine crackle glaze network across the interior surface. A quiet record of thermal history.","KR-BW-007"),
  ("IMG_5286","Matte Black Bowl","Bowls","Wheel-thrown","For Sale",50.00,"Uncompromising matte black glaze on a well-thrown form.","KR-BW-008"),
  ("IMG_5287","Speckled Stoneware Bowl","Bowls","Wheel-thrown","For Sale",50.00,"Iron speckle through a warm, sand-toned stoneware glaze. Everyday presence, quiet endurance.","KR-BW-009"),
  ("IMG_5288","Wide Foot Bowl","Bowls","Wheel-thrown","Private Collection",None,"An unusually wide foot ring lifts this bowl with a studied architectural confidence.","KR-BW-010"),
  ("IMG_5225","Sgraffito Carved Vase","Pitchers & Jugs","Wheel-thrown","For Sale",50.00,"A pulled-lip pitcher in pale celadon with a ribbed body. Elegant in form, reliable in the pour.","KR-PJ-001"),
  ("IMG_5226","Amber Glaze Jug","Pitchers & Jugs","Wheel-thrown","For Sale",50.00,"Warm amber glaze drips slightly at the shoulder, evidence of the kiln's own decisions.","KR-PJ-002"),
  ("IMG_5270","Slip-Trailed Pitcher","Pitchers & Jugs","Wheel-thrown","For Sale",50.00,"Contrasting slip trails decorate the surface in a loose, gestural pattern. Utility elevated.","KR-PJ-003"),
  ("IMG_5271","Tall-Neck Jug","Pitchers & Jugs","Wheel-thrown","For Sale",50.00,"An elongated neck sits above a broad shoulder and tapered base.","KR-PJ-004"),
  ("IMG_5272","Stoneware Creamer","Pitchers & Jugs","Wheel-thrown","For Sale",50.00,"A small-scale pouring form with a pinched spout and looped handle.","KR-PJ-005"),
  ("IMG_5273","Rustic Lug Jug","Pitchers & Jugs","Hand-built","For Sale",50.00,"Hand-built with a confident roughness. Applied lug handle and wide mouth give it an old-world ease.","KR-PJ-006"),
  ("IMG_5252","Ash Glaze Platter","Platters & Dishes","Slab-built","For Sale",50.00,"A slab-built platter with an ash-over-iron glaze that pools in the center with painterly depth.","KR-PD-001"),
  ("IMG_5256","Oval Serving Dish","Platters & Dishes","Slab-built","For Sale",50.00,"A low-lipped oval form ideal for sharing. The glaze surface reads differently in each light.","KR-PD-002"),
  ("IMG_5257","Inlaid Clay Dish","Platters & Dishes","Slab-built","For Sale",50.00,"Dark clay inlaid into a pale body creates a graphic pattern that emerges through the glaze surface.","KR-PD-003"),
  ("IMG_5274","Wide Rim Platter","Platters & Dishes","Wheel-thrown","For Sale",50.00,"A wide, trimmed rim frames a broad flat floor. Scale gives this platter a ceremonial quality.","KR-PD-004"),
  ("IMG_5275","Landscape Platter","Platters & Dishes","Slab-built","For Sale",50.00,"Glaze application recalls a landscape horizon — the meeting of earth tones and a pale sky-wash center.","KR-PD-005"),
  ("IMG_5276","Raku Platter","Platters & Dishes","Hand-built","For Sale",50.00,"Post-firing reduction creates the silvery carbon markings across the surface of this raku platter.","KR-PD-006"),
  ("IMG_4930","Ruffled Rim Vessel","Vessels & Vases","Wheel-thrown","Private Collection",None,"A ruffled rim rises from a compressed shoulder form. Retained as the definitive example of the rim series.","KR-VV-001"),
  ("IMG_5204","Narrow Neck Vase","Vessels & Vases","Wheel-thrown","For Sale",50.00,"A tall, narrow-neck vase with an iron-saturate glaze that crawls toward the shoulder in warm browns.","KR-VV-002"),
  ("IMG_5207","Ash Glaze Vessel No. 1","Vessels & Vases","Wheel-thrown","For Sale",50.00,"First of the ash glaze vessel series. The surface catches light along every thrown line.","KR-VV-003"),
  ("IMG_5208","Ash Glaze Vessel No. 2","Vessels & Vases","Wheel-thrown","Private Collection",None,"Second in the ash glaze series, kept as a paired reference for the collection.","KR-VV-004"),
  ("IMG_5209","Ash Glaze Vessel No. 3","Vessels & Vases","Wheel-thrown","For Sale",50.00,"Third vessel in the ash series, with the strongest surface variation.","KR-VV-005"),
  ("IMG_5211","Textured Stoneware Vase","Vessels & Vases","Coil-built","Private Collection",None,"Coil-built with a deliberately rough outer texture under a thin wash of pale glaze.","KR-VV-006"),
  ("IMG_5213","Ochre Glaze Vase","Vessels & Vases","Wheel-thrown","Private Collection",None,"A deep ochre glaze sits heavily on the lower half, thinning to near-transparent at the neck.","KR-VV-007"),
  ("IMG_5217","Reduction Vessel","Vessels & Vases","Wheel-thrown","For Sale",50.00,"Reduction firing produced the bronzed metallic sheen across this vessel's upper register.","KR-VV-008"),
  ("IMG_5218","Folded Rim Celadon","Vessels & Vases","Wheel-thrown","For Sale",50.00,"A wide vessel with a folded, softly irregular rim. The celadon pools at the base with seafoam clarity.","KR-VV-009"),
  ("IMG_5219","Stacked Form Study","Sculptural Works","Hand-built","Private Collection",None,"A stacked, interlocking form study. Not intended as vessel — made to explore clay under compression.","KR-SW-001"),
  ("IMG_5223","Shard Assemblage No. 1","Sculptural Works","Hand-built","For Sale",50.00,"Fired clay shards assembled into a deliberate rupture. A meditation on breakage and reconstruction.","KR-SW-002"),
  ("IMG_5232","Coiled Tower","Sculptural Works","Coil-built","Private Collection",None,"A narrow coil-built tower with exposed coil texture and a raw top edge.","KR-SW-003"),
  ("IMG_5234","Fractured Rim Form","Sculptural Works","Wheel-thrown","For Sale",50.00,"A vessel-derived form interrupted at the rim — the clay tears away deliberately.","KR-SW-004"),
  ("IMG_5259","Clay Relief Panel","Sculptural Works","Slab-built","Private Collection",None,"A flat slab panel with impressed and carved relief. Wall-hung or freestanding on its own base.","KR-SW-005"),
  ("IMG_5262","Torqued Column","Sculptural Works","Coil-built","Private Collection",None,"A column form torqued slightly from vertical. The twist is subtle but visible in profile.","KR-SW-006"),
  ("IMG_5266","Compressed Globe","Sculptural Works","Wheel-thrown","For Sale",50.00,"A thrown sphere compressed from above — a form on the edge of collapse, held in permanent tension.","KR-SW-007"),
  ("IMG_5268","Pinched Wall Form","Sculptural Works","Hand-built","Private Collection",None,"Pinch-formed walls rise and fold into each other. The fingermarks remain as part of the surface language.","KR-SW-008"),
  ("IMG_5279","Fired Clay Fragment","Sculptural Works","Hand-built","Private Collection",None,"A large fired fragment — part vessel, part ruin. The piece records something between intention and accident.","KR-SW-009"),
]

# ── Styles ───────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1C1C1A")
HDR_FONT  = Font(name="Calibri", bold=True, color="D4A017", size=11)
FS_FILL   = PatternFill("solid", fgColor="E8F5EE")
PC_FILL   = PatternFill("solid", fgColor="FFF8E7")
THIN      = Side(style="thin", color="D3D1C7")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP      = Alignment(wrap_text=True, vertical="center")
CENTER    = Alignment(horizontal="center", vertical="center")
ROW_H     = 68
THUMB_H   = 80

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "KRPots Inventory"

# ── Headers ──────────────────────────────────────────────────────────────────
headers = ["Photo", "SKU", "Title", "Category", "Technique", "Status", "Price (USD)", "Description", "Image URL"]
ws.append(headers)
for col in range(1, len(headers) + 1):
    cell = ws.cell(1, col)
    cell.fill = HDR_FILL
    cell.font = HDR_FONT
    cell.alignment = CENTER
    cell.border = BORDER
ws.row_dimensions[1].height = 22

col_widths = [14, 14, 28, 22, 16, 18, 14, 52, 55]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Data rows ────────────────────────────────────────────────────────────────
ok = 0
missing = 0
for row_idx, (img_id, title, cat, tech, status, price, desc, sku) in enumerate(pieces, 2):
    ws.row_dimensions[row_idx].height = ROW_H
    fill = FS_FILL if status == "For Sale" else PC_FILL

    row_data = [
        "",
        sku,
        title,
        cat,
        tech,
        status,
        price if price else "N/A",
        desc,
        "{}/{}.webp".format(PROD_BASE, img_id),
    ]
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row_idx, col_idx, val)
        cell.fill = fill
        cell.border = BORDER
        cell.alignment = WRAP
        if col_idx == 7 and price:
            cell.number_format = '"$"#,##0.00'
        if col_idx == 6:
            cell.font = Font(bold=True, color="1A5C38" if status == "For Sale" else "8A6000")

    # Embed thumbnail
    webp_path = os.path.join(MEDIA, "{}.webp".format(img_id))
    if os.path.exists(webp_path):
        try:
            pil = PILImage.open(webp_path).convert("RGB")
            scale = THUMB_H / pil.height
            new_w = max(1, int(pil.width * scale))
            pil = pil.resize((new_w, THUMB_H), PILImage.LANCZOS)
            buf = io.BytesIO()
            pil.save(buf, format="PNG")
            buf.seek(0)
            xl_img = XLImage(buf)
            xl_img.anchor = "A{}".format(row_idx)
            ws.add_image(xl_img)
            ok += 1
        except Exception as e:
            ws.cell(row_idx, 1, str(e))
            missing += 1
    else:
        ws.cell(row_idx, 1, "missing")
        missing += 1

ws.freeze_panes = "B2"
ws.auto_filter.ref = "A1:{}{}".format(get_column_letter(len(headers)), len(pieces) + 1)

wb.save(OUT)
print("Saved: {}".format(OUT))
print("Rows: {}  |  Images embedded: {}  |  Missing: {}".format(len(pieces), ok, missing))
